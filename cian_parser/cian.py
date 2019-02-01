#!/usr/bin/python
# -*- coding: utf-8 -*-
import os
import time
#import shutil
#import requests
import openpyxl
#from test import get_phone
#from bs4 import BeautifulSoup
from random import choice
#from fake_useragent import UserAgent
#from datetime import datetime, timedelta
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException


class ParamsURLCian:
    """ Link params for pars combination 
    BASE URL must have cat.php? in https://www.cian.ru/.
    Don't change this!
    """
    BASE_URL = 'https://www.cian.ru/cat.php?'
    DEAL_TYPE_SALE = 'deal_type=sale' # must be only after URL ( BASE_URL + DEAL_TYPE_SALE or BUY )
    ENGINE = '&engine_version=2'
    OFFER_FLAT = '&offer_type=flat'
    HOME_OWNER = '&is_by_homeowner=1'

    # BASR URL FOR WORK WITH CONTRIBUTION
    # all skan object must be protect of popertier
    URL = BASE_URL + DEAL_TYPE_SALE + ENGINE + HOME_OWNER
    
    SECOND = '&object_type%5B0%5D=1'
    NEW = '&object_type%5B0%5D=2'

    MOSKVA = '&region=1' # 1 - Moskow default
    SPB = '&region=2'
    VORONEJ = '&region=4713'
    PSKOV = '&region=4946'
    KURSK = '&region=4835'
    NIJNIY_NOVGOROD = '&region=4885'
    VELIKIY_NOVGOROD = '&region=4694'
    JAROSLAVL = '&region=5075'
    KALUGA = '&region=4780'
    TVER = '&region=176083'
    TULA = '&region=5020'
    RJAZAN = '&region=4963'
    VLADIMIR = '&region=4703'
    IVANOVO = '&region=4767'
    BRJANSK = '&region=4691'
    SMOLENSK = '&region=4987'
    OREL = '&region=175604'
    ROSTOV = '&region=4671'
    BELGOROD = '&region=4671'
    ROSTOV_NA_DONU = '&region=4959'

    ROOM_0 = '&room0=1'
    ROOM_1 = '&room1=1'
    ROOM_2 = '&room2=1'
    ROOM_3 = '&room3=1'
    ROOM_4 = '&room4=1'
    ROOM_5 = '&room5=1'
    ROOM_6 = '&room6=1'
    
    FREE_PLAN = '&room7=1'
    STUDIO = '&room9=1'

    #MIN_AREA = '&mintarea='
    #MAX_AREA = '&maxtarea='
    
    OFFER_SUBURBAN = '&offer_type=suburban'
    
    CITIES_NUMS = {
        '0': 'Москва',
        '1': 'Петербург',
        '2': 'Воронеж',
        '3': 'Псков',
        '4': 'Курск',
        '5': 'Нижний Новгород',
        '6': 'Великий Новгород',
        '7': 'Ярославль',
        '8': 'Калуга',
        '9': 'Тверь',
        '10': 'Тула',
        '11': 'Рязань',
        '12': 'Владимир',
        '13': 'Иванво',
        '14': 'Брянск',
        '15': 'Смоленск',
        '16': 'Орёл',
        '17': 'Ростов',
        '18': 'Белогород',
        '19': 'Ростов На Дону'
    }

    CITIES = {
        'Москва': MOSKVA,
        'Петербург': SPB,
        'Воронеж': VORONEJ,
        'Псков': PSKOV,
        'Курск': KURSK,
        'Нижний Новгород': NIJNIY_NOVGOROD,
        'Великий Новгород': VELIKIY_NOVGOROD,
        'Ярославль': JAROSLAVL,
        'Калуга': KALUGA,
        'Тверь': TVER,
        'Тула': TULA,
        'Рязань': RJAZAN,
        'Владимир': VLADIMIR,
        'Иванво': IVANOVO,
        'Брянск': BRJANSK,
        'Смоленск': SMOLENSK,
        'Орёл': OREL,
        'Ростов': ROSTOV,
        'Белогород': BELGOROD,
        'Ростов На Дону': ROSTOV_NA_DONU,
    }

    SEARCH_PARAMS_NUMS = {
        '0': 'всё',
        '1': 'комнаты',
        '2': 'однокомнатные квартиры',
        '3': 'двухкомнатные квартиры',
        '4': 'трехкомнатные квартире',
        '5': 'четырехкомнатные квартире',
        '6': 'пятикомнатные квартире',
        '7': 'шестикомнатные квартире',
        '8': 'квартиры студии',
        '9': 'свободная планировка',
        '10': 'дома',
    }
    
    SEARCH_PARAMS = {
        '':'',
        'комнаты': ROOM_0,
        'однокомнатные квартиры': ROOM_1,
        'двухкомнатные квартиры': ROOM_2,
        'трехкомнатные квартире': ROOM_3,
        'четырехкомнатные квартире': ROOM_4,
        'пятикомнатные квартире': ROOM_5,
        'шестикомнатные квартире': ROOM_6,
        'квартиры студии': STUDIO,
        'свободная планировка': FREE_PLAN,
        'дома': OFFER_SUBURBAN,
    }

    USE_PROPERTY = {
        '2': SECOND, 
        '1': NEW,
        '0': '',
    }

class BaseCianProcessor(ParamsURLCian):
    
    result = []
    
    file_name = ''
    file = ''
    driver = webdriver.Chrome(executable_path='/usr/local/bin/chromedriver')
    url = ''


    def file_save_xlsx(self):
        
        workbook = openpyxl.load_workbook(self.file)
        sheet = workbook.active
        
        i=1
        for data in self.result: 
            c1 = sheet.cell(row=i, column=1)
            c1.value = str(i)

            c2 = sheet.cell(row=i, column=2)
            c2.value = data['title']
            
            c3 = sheet.cell(row=i, column=3)
            c3.value = data['price']
            
            c4 = sheet.cell(row=i, column=4)
            c4.value = data['address']
            
            c5 = sheet.cell(row=i, column=5)
            c5.value = data['phone']
            
            c6 = sheet.cell(row=i, column=6)
            c6.value = data['time_data']

            c7 = sheet.cell(row=i, column=7)
            c7.value = data['url']
            
            i+=1

            for col in sheet.columns:
                max_length = 0
                column = col[0].column 
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                sheet.column_dimensions[column].width = adjusted_width
            
            workbook.save(self.file)
        


    def get_page_data(self, url):
        
        self.driver.get(url)
        
        html = self.driver.find_element_by_class_name("_93444fe79c-serp--2cnUa")
        child = html.find_elements_by_class_name("_93444fe79c-card--2Jgih")
        
        directory = os.path.expanduser("~/Desktop/cian_parser/")
        
        if not os.path.exists(directory):
            os.makedirs(directory)
    
        self.file = directory + self.file_name + '.xlsx'
        
        book = openpyxl.Workbook(self.file)
        book.save(self.file)
        
        for t in child:
            phone = self.driver.execute_script(open("reader.js").read(), t)
           
            try:
                price = t.find_element_by_class_name('c6e8ba5398-header--no6qJ').text
            
            except NoSuchElementException:
                price = t.find_element_by_class_name('c6e8ba5398-header--6WXYW').text
            
            try:
                title = t.find_element_by_class_name('c6e8ba5398-subtitle--MnTqq').text
            
            except NoSuchElementException:
                title = t.find_element_by_class_name('c6e8ba5398-header--1_m0_').text

            try:
                address = t.find_element_by_class_name('c6e8ba5398-address-links--1I9u5').text
            
            except NoSuchElementException:
                address = t.find_element_by_class_name('c6e8ba5398-address-links--1I9u5').text
            
            #try:
            url = t.find_element_by_class_name('c6e8ba5398-header--1_m0_').get_attribute("href")

            #except NoSuchElementException:

            time_data = t.find_element_by_class_name('c6e8ba5398-timeLabel--2NKak').text
        
            
            data = {
                'title': title,
                'price': price,
                'address': address,
                'phone': phone,
                'time_data': time_data,
                'url':  url,
            }

            self.result.append(data)


class CianProcessor(BaseCianProcessor):
    
    def get_driver(self, url):
        self.get_page_data(url)

    
    def inp(self):

        print('города: ')
        for k, v in self.CITIES_NUMS.items():
            print(k, ' - ', v)
        
        city = str(input('номер городa: '))
        city = self.CITIES_NUMS[city]
        
        print('номер категории: ')
        for k, v in self.SEARCH_PARAMS_NUMS.items():
            print(k, ' - ', v)
        
        param = str(input('параметр поиска: '))
        param = self.SEARCH_PARAMS_NUMS[param]
        if param == 'всё':
            param = ''

        new = str(input('выберите статус:\n0 - общее,\n1 - новостройки,\n2 - втроичная недвижимость\n'))
        amount = int(input('глубина поиска в страницах: '))
        
        self.file_name = city + '_' + param
        
        try:
            
            self.URL += self.CITIES[city]
            
            self.URL += self.SEARCH_PARAMS[param]
            
            self.URL += self.USE_PROPERTY[new]

            
            
            for i in range(amount):
                page = '&p=' + str(i+1)
                print('\nсканирование адреса: ', self.URL + page)
                self.get_driver(self.URL + page)
                print('====  успешно  ====\n')
                

            self.file_save_xlsx()
            self.driver.close()
            self.driver.quit()
        
        except KeyError:
            print('вероятно введены неверные данные')
    
    
a = CianProcessor()
a.inp()